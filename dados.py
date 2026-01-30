import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from difflib import get_close_matches
from logging_config import get_logger
from exceptions import ErroNegocio, ErroTecnico
from config import obter_dados_config


logger = get_logger(__name__)


class Dados:
    def __init__(self, arqPlanilha, sede):
        self.arqPlanilha = Path(arqPlanilha)
        self.config = obter_dados_config()

        if sede == 'Matriz':
            self.base_df = pd.read_excel(self.config['base_matriz'], usecols=["ResponsávelFinanceiro", "CPF"])
        else:
            self.base_df = pd.read_excel(self.config['base_filial'], usecols=["ResponsávelFinanceiro", "CPF"])
        

    def encontrar_melhor_match(self, nome):
        """
        Recebe um nome e retorna uma tupla contendo:
        - o nome da base de dados que mais se aproxima do nome informado;
        - o CPF associado a esse nome.
        Caso nenhum nome seja encontrado, retorna (None, None).
        """
        # Extrai a lista de nomes da coluna "ResponsávelFinanceiro".
        lista_nomes = self.base_df["ResponsávelFinanceiro"].tolist()
        
        # Usa get_close_matches para encontrar o nome mais próximo.
        # n=1: retorna apenas a melhor correspondência.
        # cutoff=0.0: nenhum corte de similaridade; ajuste se necessário.
        correspondencias = get_close_matches(nome, lista_nomes, n=1, cutoff=0.85)
        
        if correspondencias:
            melhor_nome = correspondencias[0]
            # Recupera o CPF associado a esse nome.
            cpf = self.base_df.loc[self.base_df.iloc[:, 0] == melhor_nome].iloc[0, 1]
        else:
            melhor_nome = None
            cpf = None

        return melhor_nome, cpf
    
    
    def formata_planilha(self, arqPlanilha):
        try:
            wb = load_workbook(arqPlanilha)
        except Exception as e:
            logger.error(f'Falha ao abrir a planilha {arqPlanilha}\n{e}')
            raise ErroTecnico(f'Falha ao abrir a planilha {arqPlanilha}\n{e}')


        if not 'dados' in wb.sheetnames:
            self.dados_origem = pd.read_excel(arqPlanilha, 'dados_origem', header=1, skipfooter=1)

            resultado = self.dados_origem["ResponsávelFinanceiro"].apply(
                lambda nome: pd.Series(self.encontrar_melhor_match(nome), index=["ResponsávelFinanceiro", "CPF"])
            )
            self.dados_destino = self.dados_origem.drop(columns=["ResponsávelFinanceiro", "CPF"]).join(resultado)

            dados_faltantes = self.dados_destino['ResponsávelFinanceiro'].isna()
            clientes_novos = pd.DataFrame()
            if dados_faltantes.any():
                clientes_novos = self.dados_origem.loc[dados_faltantes]

                logger.info('Clientes não cadastrados encontrados')
                logger.info('\n' + '\n'.join(clientes_novos['ResponsávelFinanceiro'].to_list()))

            with pd.ExcelWriter(arqPlanilha, mode="a", engine="openpyxl", if_sheet_exists="replace") as writer:
                self.dados_destino.to_excel(writer, sheet_name="dados", index=False, startrow=1)
                if not clientes_novos.empty:
                    clientes_novos.to_excel(writer, sheet_name="clientes_novos", index=False)


    def obter_dados(self, a_fazer=True):
        try:
            wb = load_workbook(self.arqPlanilha)
        except Exception as e:
            logger.error(f'Falha ao abrir a planilha {self.arqPlanilha}\n{e}')
            raise ErroTecnico(f'Falha ao abrir a planilha {self.arqPlanilha}\n{e}')

        if 'dados' in wb.sheetnames:
            self.dados = pd.read_excel(self.arqPlanilha, 'dados', header=1)#, skipfooter=1)

            if 'Notas' not in self.dados.columns:
                self.dados['Notas'] = None
            
            self.dados['Aluno'] = self.dados['Aluno'].apply(lambda i: i.split()[0])

            self.dados.loc[self.dados['Turma'].str.contains('Y1|Y2|Year'), 'Acumulador'] = '2'
            self.dados['Acumulador'] = self.dados['Acumulador'].fillna('1')

            self.dados['Mensalidade'] = self.dados['Mensalidade'].apply(lambda x: '{:0.2f}'.format(x).replace('.',','))
            self.dados['ValorTotal'] = self.dados['ValorTotal'].apply(lambda x: '{:0.2f}'.format(x).replace('.',','))
            self.dados['Alimentação'] = self.dados['Alimentação'].apply(lambda x: '{:0.2f}'.format(x).replace('.',','))

            return self.dados[self.dados['Notas'].isna()] if a_fazer else self.dados
        else:
            logger.error("A aba 'dados' não existe na planilha")
            raise ErroNegocio(
                "A aba 'dados' não existe na planilha"
            )
    

    def registra_numero_notas(self, index_df, num_nota):
        if not isinstance(index_df, int):
            raise ErroNegocio("index_df deve ser um inteiro")
        if not isinstance(num_nota, int):
            raise ErroNegocio("num_nota deve ser um inteiro")
            
        try:
            wb = load_workbook(self.arqPlanilha)
            sheet = wb['dados']
        except Exception as e:
            logger.error(f'Falha ao abrir a planilha {self.arqPlanilha}\n{e}')
            raise ErroTecnico(f'Falha ao abrir a planilha {self.arqPlanilha}\n{e}')

        # Adicionar a coluna 'Status' se não existir
        if 'Notas' not in [celula.value for celula in sheet[2]]:
            sheet.cell(row=2, column=sheet.max_column + 1).value = 'Notas'

        notas_col_index = [celula.value for celula in sheet[2]].index('Notas') + 1
        sheet.cell(row=index_df+3, column=notas_col_index).value = num_nota

        try:
            wb.save(self.arqPlanilha)
        except Exception as e:
            logger.error(f'Falha ao salvar a planilha {self.arqPlanilha}\n{e}')
            raise ErroTecnico(f'Falha ao salvar a planilha {self.arqPlanilha}\n{e}')



if __name__ == '__main__':
    arquivo_planilha = r"C:\Users\novoa\OneDrive\Área de Trabalho\notas_MB\planilhas\zona_sul\escola_canadenseZS_nov25\Numeração de Boletos_Zona Sul_2025_NOVEMBRO.xlsx"
    dados = Dados(arquivo_planilha, 'Zona Sul')
    df = dados.obter_dados()#a_fazer=False)
    print(df)
